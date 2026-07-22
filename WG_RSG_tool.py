#!/usr/bin/env python3
"""
Randomize 40 samples onto an 8x5 plate (A01-H05) and color by group.

Input CSV expected columns:
- Sample_ID
- Group

Optional:
- Well (ignored if present; wells are reassigned)
- any extra columns are preserved in the output sheets

Output Excel workbook contains:
- Original_Sample_List: the original input order
- Sample_List: randomized samples with assigned wells
- Plate_Map: 8x5 plate layout with Sample_ID only
- Legend: group-to-color mapping

This version:
- Works on Windows in VS Code
- Accepts input/output paths via command-line arguments (with sensible defaults)
- Automatically generates distinct colors for any number of groups
- Keeps the 8x5 plate only
- Removes Group text from Plate_Map, while still using group colors

Usage:
    python WG_RSG_tool.py
    python WG_RSG_tool.py --input data/Well_template.csv --output data/plate_map.xlsx --seed 42
"""

import argparse
import colorsys
import random
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Default paths (relative to this script's directory)
_HERE = Path(__file__).parent
_DEFAULT_INPUT  = _HERE.parent / "datasheet" / "Well_template.csv"
_DEFAULT_OUTPUT = _HERE.parent / "datasheet" / "randomized_plate_map_woGroupInfo.xlsx"

# Fixed 8 x 5 plate layout
ROWS = list("ABCDEFGH")
COLS = [1, 2, 3, 4, 5]


# -----------------------------
# Helpers
# -----------------------------
def load_samples(csv_path):
    df = pd.read_csv(csv_path)

    required = {"Sample_ID", "Group"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError("Input CSV is missing required columns: {}".format(sorted(missing)))

    df = df.copy()

    # Ignore any existing Well column because wells are reassigned
    if "Well" in df.columns:
        df = df.drop(columns=["Well"])

    # Basic cleanup
    df["Sample_ID"] = df["Sample_ID"].astype(str).str.strip()
    df["Group"] = df["Group"].astype(str).str.strip()

    # Remove empty rows
    df = df[df["Sample_ID"].ne("") & df["Group"].ne("")].reset_index(drop=True)
    return df


def build_well_list(rows, cols):
    return ["{}{:02d}".format(r, c) for r in rows for c in cols]


def assign_randomized_wells(df, seed, rows, cols):
    n_wells = len(rows) * len(cols)
    if len(df) > n_wells:
        raise ValueError("Too many samples ({}) for plate size ({} wells).".format(len(df), n_wells))

    rng = random.Random(seed)

    shuffled_idx = list(df.index)
    rng.shuffle(shuffled_idx)
    randomized = df.loc[shuffled_idx].reset_index(drop=True).copy()

    wells = build_well_list(rows, cols)

    if len(randomized) < len(wells):
        blank_count = len(wells) - len(randomized)
        blank_df = pd.DataFrame({col: [""] * blank_count for col in df.columns})
        randomized = pd.concat([randomized, blank_df], ignore_index=True)

    randomized["Well"] = wells[: len(randomized)]
    return randomized


def group_color_map(groups, seed=42):
    """
    Generate a distinct pastel-like hex color for each unique group.
    Works for any number of groups.
    """
    unique_groups = list(dict.fromkeys(groups))
    n = len(unique_groups)

    if n == 0:
        return {}

    hues = [i / n for i in range(n)]
    rng = random.Random(seed)
    rng.shuffle(hues)

    mapping = {}
    for grp, h in zip(unique_groups, hues):
        r, g, b = colorsys.hls_to_rgb(h, 0.82, 0.45)
        mapping[grp] = "{:02X}{:02X}{:02X}".format(int(r * 255), int(g * 255), int(b * 255))

    return mapping


def add_border(cell):
    thin = Side(style="thin", color="808080")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="404040")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    add_border(cell)


def style_body(cell, fill_color=None):
    if fill_color:
        cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    add_border(cell)


def make_plate_matrix(df, rows, cols):
    lookup = {
        row["Well"]: {
            "Sample_ID": row["Sample_ID"],
            "Group": row["Group"],
        }
        for _, row in df.iterrows()
        if str(row["Well"]).strip() != ""
    }

    matrix = pd.DataFrame(index=rows, columns=["{:02d}".format(c) for c in cols])

    for r in rows:
        for c in cols:
            well = "{}{:02d}".format(r, c)
            sample = lookup.get(well, {}).get("Sample_ID", "")
            if sample:
                matrix.loc[r, "{:02d}".format(c)] = sample
            else:
                matrix.loc[r, "{:02d}".format(c)] = ""

    return matrix


def write_excel(original_df, randomized_df, output_xlsx, seed, rows, cols):
    wb = Workbook()
    wb.remove(wb.active)

    group_map = group_color_map(randomized_df.loc[randomized_df["Group"].ne(""), "Group"].tolist(), seed=seed)

    # -----------------
    # Original_Sample_List sheet
    # -----------------
    ws0 = wb.create_sheet("Original_Sample_List")
    extra_cols0 = [c for c in original_df.columns if c not in {"Well", "Sample_ID", "Group"}]
    out_cols0 = ["Sample_ID", "Group"] + extra_cols0

    for col_idx, col_name in enumerate(out_cols0, start=1):
        cell = ws0.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    for row_idx, (_, row) in enumerate(original_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(out_cols0, start=1):
            value = row[col_name] if col_name in row.index else ""
            cell = ws0.cell(row=row_idx, column=col_idx, value=value)

            if col_name == "Group" and str(value).strip() != "":
                fill = group_map.get(str(value).strip(), "FFFFFF")
                style_body(cell, fill)
            else:
                style_body(cell)

    ws0.freeze_panes = "A2"
    for i in range(1, len(out_cols0) + 1):
        ws0.column_dimensions[get_column_letter(i)].width = 18

    # -----------------
    # Sample_List sheet
    # -----------------
    ws1 = wb.create_sheet("Sample_List")
    extra_cols = [c for c in randomized_df.columns if c not in {"Well", "Sample_ID", "Group"}]
    out_cols = ["Well", "Sample_ID", "Group"] + extra_cols

    for col_idx, col_name in enumerate(out_cols, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)

    for row_idx, (_, row) in enumerate(randomized_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(out_cols, start=1):
            value = row[col_name] if col_name in row.index else ""
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)

            if col_name == "Group" and str(value).strip() != "":
                fill = group_map.get(str(value).strip(), "FFFFFF")
                style_body(cell, fill)
            else:
                style_body(cell)

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 14
    for i in range(4, len(out_cols) + 1):
        ws1.column_dimensions[get_column_letter(i)].width = 18
    ws1.freeze_panes = "A2"

    # -----------------
    # Plate_Map sheet
    # -----------------
    ws2 = wb.create_sheet("Plate_Map")
    plate_title = "Plate Map ({}{:02d}-{}{:02d})".format(rows[0], cols[0], rows[-1], cols[-1])
    ws2["A1"] = plate_title
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols) + 1)

    ws2["A2"] = ""
    style_header(ws2["A2"])
    for col_idx, c in enumerate(cols, start=2):
        cell = ws2.cell(row=2, column=col_idx, value="{:02d}".format(c))
        style_header(cell)

    matrix = make_plate_matrix(randomized_df, rows, cols)

    for row_idx, r in enumerate(rows, start=3):
        row_cell = ws2.cell(row=row_idx, column=1, value=r)
        style_header(row_cell)

        for col_idx, c in enumerate(["{:02d}".format(x) for x in cols], start=2):
            value = matrix.loc[r, c]
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)

            fill_color = None
            if value:
                sample_row = randomized_df.loc[randomized_df["Well"] == "{}{:02d}".format(r, int(c))]
                if not sample_row.empty:
                    grp = str(sample_row.iloc[0]["Group"]).strip()
                    fill_color = group_map.get(grp)

            style_body(cell, fill_color)

    ws2.column_dimensions["A"].width = 8
    for col in range(2, len(cols) + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    for row in range(3, 3 + len(rows)):
        ws2.row_dimensions[row].height = 30
    ws2.freeze_panes = "B3"

    # -----------------
    # Legend sheet
    # -----------------
    ws3 = wb.create_sheet("Legend")
    ws3["A1"] = "Group"
    ws3["B1"] = "Color"
    style_header(ws3["A1"])
    style_header(ws3["B1"])

    for i, (grp, color) in enumerate(group_map.items(), start=2):
        a = ws3.cell(row=i, column=1, value=grp)
        b = ws3.cell(row=i, column=2, value="#" + color)
        style_body(a, color)
        style_body(b)
        b.fill = PatternFill("solid", fgColor=color)

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 14

    wb.save(output_xlsx)
    print("Saved: {}".format(output_xlsx))
    print("Randomization seed: {}".format(seed))
    print("Sheets created: Original_Sample_List, Sample_List, Plate_Map, Legend")


def parse_args():
    parser = argparse.ArgumentParser(description="Randomize samples onto a plate and export to Excel.")
    parser.add_argument("--input",  type=Path, default=_DEFAULT_INPUT,  help="Path to input CSV (default: %(default)s)")
    parser.add_argument("--output", type=Path, default=_DEFAULT_OUTPUT, help="Path to output XLSX (default: %(default)s)")
    parser.add_argument("--seed",   type=int,  default=42,              help="Random seed (default: %(default)s)")
    return parser.parse_args()


def main():
    args = parse_args()

    if not args.input.exists():
        print("Error: cannot find input file: {}".format(args.input), file=sys.stderr)
        sys.exit(1)

    original_df = load_samples(args.input)
    randomized_df = assign_randomized_wells(original_df, args.seed, ROWS, COLS)
    write_excel(original_df, randomized_df, args.output, args.seed, ROWS, COLS)


if __name__ == "__main__":
    main()

